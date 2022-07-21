# Tutorial 1 = print
# print("Anna Hartl")
# print('o----')
# print(' ||||')
# print('*' * 1)
# print('*' * 2)
# print('*' * 3)

# Tutorial 2 = Variables
# age = 20
# full_name = 'John Smith'
# is_new_patient = True
# print(full_name)

# Tutorial 3 = input
# name = input('What is your name? ')
# color = input('What is your favourite color? ')
#
# print(name+ ' likes '+color)

# Tutorial 4 = Type conversion
# birth_year = input('Birth Year: ')
# print(type(birth_year))
# age = 2022 - int(birth_year)
# print(age)
# print(type(age))
# weight_punds = input('Your weight in poungs: ')
# weight_kilo = float(weight_punds) * 0.453592
#
# print('Your wweight in kilo: '+ str(weight_kilo))

# Tutorial 5 = Strings

# course = 'Python for beginners'
# another = course[:]
# print(another)

# name = 'Jennifer'
# print(name[1:-1])

# Tutorial 6 = formatted Strings
# first_name = 'John'
# last_name = 'Smith'
#
# message = first_name + '[' + last_name + '] is a coder'
# msg = f'{first_name} [ {last_name} ] is a coder'
# print(msg)
# print(message)

# Tutorial 7 = String Methods
#
# course = 'Python for Beginners'
# # Method wenn mit Punkt angesprochen wird
# print(course.upper())
# print(course.lower())
# print(course.title())
# print(course.find('for'))
# print(course.replace('P', 'J'))
# print('python' in course)
# print(course)

# Tutorial 8 = arithmetic Operations
# print(10 ** 3)
# x = 10
# x **= 3
# print(x)

# Tutorial 9 = Operator Precedence
# x= (2+ 3) * 10 -3
# print(x)

# Tutorial 10 = Math functions
# import math
#
# x = -2.9
# print(abs(x))
# print(math.floor(2.9))

# Tutorial 11 = if statements
# is_hot = False
# is_cold = False
# if is_hot:
#     print("It's a hot day")
#     print("Drink plenty of water")
#
# elif is_cold:
#     print("It's a cold day")
#     print("Wear warm clothes")
#
# else:
#     print("It's a lovely day")
#
# price = 1000000
# has_good_credit = True
#
# if has_good_credit:
#     downpayment = price * 0.1
# else:
#     downpayment = price * 0.2
#
# print(f"Downpayment: {downpayment} $")

# Tutorial 12 = Logical Operators
#
# has_high_income = False
# has_good_credit = True
# has_criminal_record = True
#
# if has_high_income or has_good_credit:
#     print("Eligible for loan")
#
# if has_good_credit and not has_criminal_record:
#     print("Eligible for loan")

# Tutorial 13 = Comparison Operators
# temperature = 35
#
# if temperature > 30:
#     print("It's a hot day")
# else:
#     print("it's not a hot day")
#
# name = input("Your name: ")
#
# if len(name) < 3:
#     print("Name musst be at least 3 characters long")
# elif len(name) > 50:
#     print("Name has a maximum of 50 character")
# else:
#     print("Name looks good")

# Tutorial 13 = Project Weight converter
#
# weight = float(input("Weight: "))
# char_conversion = input("(L)bs or (K)ilo: ")
#
#
# if char_conversion.lower() == 'l':
#     converted_weight = weight * 0.453592
#     print(f"You are {round(converted_weight, 2)} kilos")
# else:
#     converted_weight = weight * 2.20462
#     print(f"You are {round(converted_weight, 2)} pounds")

# Tutorial 14 = While loops
#
# i = 1
#
# while i <= 5:
#     print('*'*i)
#     i+= 1
# print("done")

# Tutorial 15 = Project Guessing game
#
# secret_number = 9
# guess_count = 0
# guess_limit = 3
#
# while guess_count < guess_limit :
#     guess = int(input("Guess: "))
#     guess_count += 1
#     if guess == secret_number:
#         print("You won!")
#         break
# else:
#     print("Sorry, you failed")

# Tutorial 16 = Car game
#
# user_text = ""
# car_started = False
#
# while True:
#
#     user_text = input('>').lower()
#
#     if user_text == "help":
#         print('start - to start the car\nstop - to stop the car\nquit - to exit')
#     elif user_text == "start":
#         if car_started:
#             print("Car is already started")
#         else:
#             print('Car started ... ready to go')
#         car_started = True
#     elif user_text == "stop":
#         if not car_started:
#             print("Car is already stopped!")
#         else:
#             print("Car stopped")
#         car_started = False
#     elif user_text == "quit":
#         break
#     else:
#         print("I don't understand")

# Tutorial 17 = For loops

# for item in range(5,10):
#     print(item)
# prices = [10,20,30]
#
# total_price = 0
# for shopping_item in prices:
#     total_price += shopping_item
#
# print(total_price)

# Tutorial 17 = nested loops
#
# for x in range(4):
#     for y in range(3):
#         print(f"({x}, {y})")
#
# numbers = [1,1,1,1,5]
# st = ""
# for item in numbers:
#     for x in range(item):
#         st += "x"
#     print(st)
#     st = ""

# Tutorial 18 = Lists
#
# names = ["John", "Sarah", "Mosh"]
# names[0] = "jon"
# print(names)
#
# numbers = [2,5,10,7,2]
# biggest_num = 0
# for num in numbers:
#     if num > biggest_num:
#         biggest_num = num
#
# print(biggest_num)

#Tutorial 19 = 2D Lists
#
# matrix = [
#     [1,2,3],
#     [1,2,3],
#     [1,2,3]
# ]
#
# matrix[1][0] = 20
#
# print(matrix[1][0])

# Turorial 20 = List Methods

# numbers = [1,2, 4, 4, 4,4, 2]
#
# numbers.append(20)
# numbers.pop()
# numbers.sort()
# numbers.reverse()
#
# for item in numbers:
#     while numbers.count(item) > 1:
#         numbers.remove(item)
#
# print(numbers)

# Tutorial 21 = Tuples
# numbers = (1,2,4)

# Tutorial 22 = unpacking
#
# coordinates = (1,2,3)
#
# x = coordinates[0]
# y = coordinates[1]
# z = coordinates[2]
#
# x ,y , z = coordinates

#  Tutorial 23 = Dictonaries
#
# customer = {
#     "name": "John Smith",
#     "age": 30,
#     "is_varified": True
# }
#
# print(customer.get("name", "test"))
#
# phone_numbers = {
#     "0": "Zero",
#     "1": "One",
#     "2": "Two",
#     "3": "Tree",
#     "4": "Four",
#     "5": "Five",
#     "6": "Six",
#     "7": "Seven",
#     "8": "Eight",
#     "9": "Nine"
# }
#
# user_input = input("Phone: ")
# phone_number_words = ""
#
# for item in user_input:
#     phone_number_words += phone_numbers.get(item, "!") + " "
#
# print(phone_number_words)
#
# message = input("> ")
# words = message.split(" ")
# emojis = {
#     ":)": "happy face",
#     ":(": "sad face"
# }
#
# output = ""
#
# for word in words:
#     output += emojis.get(word,word) + " "
#
# print(output)

# Tutorial 24 = Functions
#
# def greet_user():
#     print("Hi there!")
#     print("Welcome abroad")
#
#
# greet_user()

# Tutorial 25 = Parameter

# def greet_user(first_name, last_name):
#     print(f"Hi {first_name} {last_name}!")
#     print("Welcome abroad")
#
#
# greet_user("John", "Smith")

# Tutorial 26 = Keyword Arguments
#
# def greet_user(first_name ="there", last_name = ""):
#     print(f"Hi {first_name} {last_name}!")
#     print("Welcome abroad")
#
#
# greet_user(first_name="John", last_name="Smith")
# greet_user("Mary")
# greet_user()

# Tutorial 27 = return Statement
#
# def square(number):
#     return number * number
#
#
# print(square(3))

# Tutorial 28 = creating a reuseable function

# def emoji_converter(msg):
#     words = msg.split(" ")
#     emojis = {
#         ":)": "happy face",
#         ":(": "sad face"
#     }
#     output = ""
#     for word in words:
#         output += emojis.get(word,word) + " "
#     return output
#
#
# message = input("> ")
# print(emoji_converter(message))

# Tutorial 29 = Exceptions
#
# try:
#     age = int(input("Age: "))
#     income = 2000
#     risk  = income / age
#     print(age)
# except ZeroDivisionError:
#     print("Age cannot be zero")
# except ValueError:
#     print("Invalid value")

# Tutorial 30 = Comments

# Tutorial 31 = Classes
# Tutorial 32 = Constructor
#
# class Point:
#     def __init__(self, x, y):
#         self.x = x
#         self.y = y
#
#     def move(self):
#         print("move")
#
#     def draw(self):
#         print("draw")
#
#
# class Person:
#     def __init__(self, name):
#         self.name = name
#
#     def talk(self):
#         print(f"Hello I am {self.name}")
#
#
# p1 = Person("John")
# p1.talk()

# Tutorial 33 = Inheritance dry = dont repeat yourself

# class Mammal:
#
#     def walk(self):
#         print("walk")
#
# class Dog(Mammal):
#     def bark(self):
#         print("Wuff wuff")
#
# class Cat(Mammal):
#     def make_noise(self):
#         print("Miau Miau")
#
#
# dog1 = Dog()
# dog1.bark()
#
# cat1 = Cat()
# cat1.make_noise()

# Tutorial 34 = Modules
#
# from converters import emoji_converter
#
# print(emoji_converter(":)"))
#
# import utils
#
# numbers = [2,5,10,7,2]
#
# print(utils.find_max(numbers))

# Tutorial 35 = Packages
#
# from ecommerce import shipping
#
# shipping.calculate_shipping()

# Tutorial 36 = Generating Random Values

import random
#
# memebers = ["J", "M", "A"]
#
# for i in range(3):
#     print(random.choice(memebers))
#
#
#
# class Dice:
#     def roll(self):
#         first_roll = random.randint(1,6)
#         second_roll = random.randint(1,6)
#         return first_roll, second_roll
#
#
# dice = Dice()
# print(dice.roll())

# Tutorial 37 = Files and Diretories

# from pathlib import Path
#
# path = Path()
#
# for file in path.glob("*.py"):
#     print(file)

# Tutorial 38 = Pypi and Pip
#https://pypi.org/

# Tutorial 39 = Project Excel Spreadsheet

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def prozess_workbook(filename):
    wb = xl.load_workbook(filename)

    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row + 1):
        corrected_price = sheet.cell(row, 3).value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)

